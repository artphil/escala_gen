'''
Programa de geracao automatica de escala de postos de servico
Gerador da planilha Excel
autor: Arthur Phillip Silva
'''

from openpyxl import Workbook, formula, drawing
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell.cell import Cell
from datetime import date

import os

path_qrcode = r'img/qrcode.png'

# Cria a planilha excel
def Gen_xls(tabela, path_output):

	# Definindo a planilha
	wb = Workbook()
	ws = wb.active
	ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
	ws.page_setup.paperSize = ws.PAPERSIZE_A4
	ws.page_setup.fitToPage = True
	ws.print_area = 'A1:AK30'

	# Titulo da aba
	ws.title = 'escala-posto'

	# Preenchendo com dados da tabela
	i = 0
	for linha in tabela:
		i += 1
		ws.append(linha)

	# ws.append(('','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','@artphil'))

	# Aplicando QRCODE
	qrcode = drawing.image.Image(path_qrcode)
	qrcode.width = 50.0
	qrcode.height = 50.0
	# qrcode.anchor(ws.cell('AD20'))
	# ws.add_image(qrcode)
	ws.add_image(qrcode, 'AD23')

	# Inserindo Data
	ws["A30"].value = date.today().strftime("%d/%m/%y")

	# Aplica merge e formata as celulas do titulo
	ws.merge_cells("A1:AK1")
	ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
	ws['A1'].font = Font(bold=True)

	# Aplica formatacao as demais celulas
	for linha in ws['A2:AK23']:
		# Dimensionamento
		ws.row_dimensions[linha[0].row].height = 20.0
		# Linhas de postos
		cinza = linha[0].row % 2 == 0 and linha[0].value
		for celula in linha:
			# Orientacao do texto
			celula.alignment = Alignment(horizontal='center', vertical='center')
			# Titulos e nomes em negrito
			if celula.col_idx < 3 or celula.row < 4:
				celula.font = Font(bold=True)
			# Linhas cinzas intercaladas
			if cinza:
				celula.fill = PatternFill("solid", fgColor="F5F5F5")
			# celulas de folga em negativo
			if celula.value == 'F':
				celula.fill = PatternFill("solid", fgColor="000000")
				celula.font = Font(color='FFFFFF')

	# Definindo largura das células
	ws.row_dimensions[1].height = 30
	ws.column_dimensions['A'].width = 20.0
	ws.column_dimensions['B'].width = 5.0
	for col in range(3,40):
	     ws.column_dimensions[get_column_letter(col)].width = 4.0


	# Salva a planilha
	filename = tabela[0][0].replace(" ", "")
	
	xls_file = os.path.join(path_output, f'{filename}.xlsx')
	
	wb.save(filename=xls_file)

	print(f'Arquivo {xls_file} gravado.')
