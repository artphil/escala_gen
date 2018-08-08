from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.utils import column_index_from_string

import random
import numpy as np

def imprime(tabela):
	# Definindo a planilha
	wb = Workbook()
	ws = wb.active
	ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
	ws.page_setup.paperSize = ws.PAPERSIZE_A4
	ws.page_setup.fitToPage = True
	ws.print_area = 'A1:AG15'

	# Titulo da aba
	ws.title = 'Escala'

	# Preenchendo com dados da tabela
	i = 0;
	for linha in tabela:
		ws.append(linha)

	# Aplica merge e formata as celulas do titulo
	ws.merge_cells("A1:AG1")
	ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
	ws['A1'].font = Font(bold=True)

	# Aplica formatacao as demais celulas
	for linha in ws['A2:AG15']:
		ws.row_dimensions[linha[0].row].height = 20.0
		for celula in linha:
			celula.alignment = Alignment(horizontal='center', vertical='center')
			if celula.col_idx < 3 or celula.row < 4:
				celula.font = Font(bold=True)
			if celula.value == 'F':
				celula.fill = PatternFill("solid", fgColor="000000")
				celula.font = Font(color='FFFFFF')

	# Definindo largura das células das sequência
	ws.row_dimensions[ws['A1'].row].height = 30.0
	for col in ws['B:AG']:
	     ws.column_dimensions[col[0].column].width = 4.0

	# Salvando
	nome_arquivo = tabela[0][0]+'.xlsx'
	wb.save(filename=nome_arquivo)


def le_dados():
	data = {}

	data['estacao'] = input("nome estacao: ")
	data['dia_inicio'] = int(input("dia da mudança: "))
	data['func_num'] = int(input("numero de funcionarios: "))
	data['mes'] = input("mes: ")

	func_nomes = []
	func_Ps = []
	for i in range(data['func_num']):
		func_nomes.append(input("nome: "))
		func_Ps.append(int(input("P: ")))

	data['func_nomes'] = func_nomes
	data['func_Ps'] = func_Ps

	return data

def gera_tabela():
	escala = []

	dias_mes = {
	'Janeiro' : 31,
	'Fevereiro' : 28,
	'Marco' : 31,
	'Abril' : 30,
	'Maio' : 31,
	'Junho' : 30,
	'Julho' : 31,
	'Agosto' : 31,
	'Setembro' : 30,
	'Outubro' : 31,
	'Novembro' : 30,
	'Dezembro' : 31
	}

	semana = ["D", "S", "T", "Q", "Q", "S", "S", "D"]

	data = {}
	data['estacao'] = "Central"
	data['dia_inicio'] = 3
	data['mes'] = 'Julho'
	data['func_num'] = 6
	data['func_nomes'] = ['Artphil', 'D.Maia', 'Waguim', 'Zeze', 'Tata', 'Tretas']
	data['func_Ps'] = [1, 4, 10, 2, 5, 11]

	data['folgas'] = [
	[0,5,10,11,15,20,25,26],
	[1,5,10,11,16,20,25,26],
	[0,4,10,11,15,20,24,25],
	[2,6,12,16,21,22,27],
	[2,6,12,15,21,22,27],
	[2,5,13,16,21,22,27]]
	data['postos'] = [2,1]

	# data = le_dados()

	escala.append(["Escala ASO1 - " + data['estacao'] + " - " + data['mes'], ""])

	lista_dias = ["Dias", "P"]
	for d in range(dias_mes[data['mes']]):
		lista_dias.append((data['dia_inicio']+d-1)%dias_mes[data['mes']]+1)
	escala.append(lista_dias)

	lista_sem = ["", ""]
	for d in range(dias_mes[data['mes']]):
		lista_sem.append(semana[d%7])
	escala.append(lista_sem)

	distrib = preenche(data['func_num'], dias_mes[data['mes']], data['postos'], data['folgas'])
	postos = []
	c = 0
	for f in range(data['func_num']):
		p = []
		p.append(data['func_nomes'][f])
		p.append(data['func_Ps'][f])
		for i in range(dias_mes[data['mes']]):
			if distrib[c][i] == 1:
				p.append("F")
			elif distrib[c][i] == 2:
				p.append("B1")
			elif distrib[c][i] == 3:
				p.append("Q1")
			elif distrib[c][i] == 4:
				p.append("B2")
			elif distrib[c][i] == 5:
				p.append("Q2")
			else:
				p.append("")
		c += 1
		escala.append(p)

	imprime(escala)

	print ("Arquivo gerado no diretorio: ")

def preenche(func=6, dias=15, postos_num=[2,1], folgas=[[0,4,7,8,11,14],[1,4,7,8,11,14],[0,4,7,8,11,14],[2,3,6,9,12],[2,3,6,9,12],[2,3,6,9,12]]):
	dist_postos = np.zeros((func,dias))

	func_fixos = int(func/2)
	postos_trab = np.zeros((func,func_fixos))

	f=0
	for lista in folgas:
		for d in lista:
			dist_postos[f][d] = 1
		f += 1

	postos = []
	m = 2
	for qtd in postos_num:
		for i in range(qtd):
			postos.append(i*2 + m)
		m += 1

	postos.sort()
	print (postos)
	
	for j in range(dias):
		k = j%func_fixos;
		for i in range(func_fixos):
			# print(i,j,k)
			if dist_postos[i][j] == 1:
				i += func_fixos
			dist_postos[i][j] = postos[k]
			k = (k+1)%func_fixos
	'''
	for d in range(dias):
		postos_ocp = np.zeros(func_fixos)
		num_postos = func_fixos
		f = random.randint(0,func_fixos)
		while num_postos > 0:
			print(f)
			if dist_postos[f][d] != 0:
				if dist_postos[f][d] == 1:
					f = (f+func_fixos)%func
				else:
					f = (f+1)%func_fixos
			else:
				p_min=999
				p_i = 0
				for i in range(func_fixos):
					if postos_ocp[i] == 0 and postos_trab[f][i] < p_min:
						p_min = postos_trab[f][i]
						p_i = i
				print(f, p_i)
				postos_ocp[p_i] = 1
				postos_trab[f][p_i] += 1
				dist_postos[f][d] = postos[p_i]
				num_postos -= 1
				# f = (f+1)%func_fixos
	''' '''
	for d in range(dias):
		fix = []
		res = []
		for f in range(func_fixos):
			if dist_postos[f][d] == 1:
				res.append(f+func_fixos)
			else:
				fix.append(f)

		for p in range(func_fixos):
			f_min=999
			f_i = 0
			for i in range(func_fixos):
				if postos_trab[i][p] < f_min:
					f_min = postos_trab[i][p]
					f_i = i

			postos_trab[f_i][p] += 1
			postos_trab[f_i += func_fixos][p] += 1

			if dist_postos[f_i][d] == 1:
				f_i += func_fixos

			print(f_i, p)

			dist_postos[f_i][d] = postos[p]
	'''
	print (dist_postos)
	return dist_postos

def aloca():
	pass


gera_tabela()
# preenche()
