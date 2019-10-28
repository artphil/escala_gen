'''
Programa de geracao automatica de escala de postos de servico
Gerador da planilha Excel
autor: Arthur Phillip Silva
'''

from openpyxl import Workbook, formula, drawing
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import Cell
from datetime import date

# Cria a planilha excel
def gera_xls(tabela):
	# Definindo a planilha
	wb = Workbook()
	ws = wb.active
	ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
	ws.page_setup.paperSize = ws.PAPERSIZE_A4
	ws.page_setup.fitToPage = True
	ws.print_area = 'A1:AK30'

	# Titulo da aba
	ws.title = 'escala-posto'

	# Preenchendo com dados da tabela
	i = 0;
	for linha in tabela:
		i += 1
		ws.append(linha)

	# ws.append(('','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','','@artphil'))

	# Aplicando QRCODE
	qrcode = drawing.image.Image('img/qrcode.png')
	qrcode.width = 50.0
	qrcode.height = 50.0
	# qrcode.anchor(ws.cell('AD20'))
	# ws.add_image(qrcode)
	ws.add_image(qrcode, 'AD23')

	# Inserindo Data
	ws["A30"].value = date.today().strftime("%d/%m/%y")

	# Aplica merge e formata as celulas do titulo
	ws.merge_cells("A1:AK1")
	ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
	ws['A1'].font = Font(bold=True)

	# Aplica formatacao as demais celulas
	for linha in ws['A2:AK23']:
		# Dimensionamento
		ws.row_dimensions[linha[0].row].height = 20.0
		# Linhas de postos
		cinza = linha[0].row % 2 == 0 and linha[0].value
		for celula in linha:
			# Orientacao do texto
			celula.alignment = Alignment(horizontal='center', vertical='center')
			# Titulos e nomes em negrito
			if celula.col_idx < 3 or celula.row < 4:
				celula.font = Font(bold=True)
			# Linhas cinzas intercaladas
			if cinza:
				celula.fill = PatternFill("solid", fgColor="F5F5F5")
			# celulas de folga em negativo
			if celula.value == 'F':
				celula.fill = PatternFill("solid", fgColor="000000")
				celula.font = Font(color='FFFFFF')

	# Definindo largura das células
	ws.row_dimensions[ws['A1'].row].height = 30
	ws.column_dimensions[ws['A1'].column].width = 20.0
	ws.column_dimensions[ws['B1'].column].width = 5.0
	for col in ws['B:AK']:
	     ws.column_dimensions[col[0].column].width = 4.0

	# Aplicando formulas de contagem de postos (nao impresso)
	ws["AM3"] = "B1"
	ws["AN3"] = "B2"
	ws["AO3"] = "B3"
	ws["AP3"] = "B4"
	ws["AQ3"] = "B5"
	ws["AR3"] = "B6"
	ws["AS3"] = "Q1"
	ws["AT3"] = "Q2"
	ws["AU3"] = "Q3"
	ws["AV3"] = "Q4"

	for i in range(4,len(tabela)+1):
		# ws["AI"+str(i)]='=CONT.SE(B'+str(i)+':AG'+str(i)+'; "B1")'
		ws["AM"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "B1")', data_type=Cell.TYPE_FORMULA)
		ws["AN"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "B2")', data_type=Cell.TYPE_FORMULA)
		ws["AO"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "B3")', data_type=Cell.TYPE_FORMULA)
		ws["AP"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "B4")', data_type=Cell.TYPE_FORMULA)
		ws["AQ"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "B5")', data_type=Cell.TYPE_FORMULA)
		ws["AR"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "B6")', data_type=Cell.TYPE_FORMULA)
		ws["AS"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "Q1")', data_type=Cell.TYPE_FORMULA)
		ws["AT"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "Q2")', data_type=Cell.TYPE_FORMULA)
		ws["AU"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "Q3")', data_type=Cell.TYPE_FORMULA)
		ws["AV"+str(i)].set_explicit_value(value='=COUNTIF(B'+str(i)+':AG'+str(i)+'; "Q4")', data_type=Cell.TYPE_FORMULA)

	# Salva a planilha
	nome_arquivo = 'planilha/'+tabela[0][0]+'.xlsx'
	wb.save(filename=nome_arquivo)
