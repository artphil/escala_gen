from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.utils import column_index_from_string

import random
import json
import sys
import numpy as np


'''                       Funcoes                              '''

def print_dic(dic):
	print (json.dumps(dic, sort_keys=True, indent=4))

# Cria a planilha excel
def gera_xls():
	# Definindo a planilha
	wb = Workbook()
	ws = wb.active
	ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
	ws.page_setup.paperSize = ws.PAPERSIZE_A4
	ws.page_setup.fitToPage = True
	ws.print_area = 'A1:AG15'

	# Titulo da aba
	ws.title = 'escala'

	# Preenchendo com dados da tabela
	i = 0;
	for linha in escala:
		ws.append(linha)

	# Aplica merge e formata as celulas do titulo
	ws.merge_cells("A1:AG1")
	ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
	ws['A1'].font = Font(bold=True)

	# Aplica formatacao as demais celulas
	for linha in ws['A2:AG15']:
		ws.row_dimensions[linha[0].row].height = 20.0
		for celula in linha:
			celula.alignment = Alignment(horizontal='center', vertical='center')
			if celula.col_idx < 3 or celula.row < 4:
				celula.font = Font(bold=True)
			if celula.value == 'F':
				celula.fill = PatternFill("solid", fgColor="000000")
				celula.font = Font(color='FFFFFF')

	# Definindo largura das células das sequência
	ws.row_dimensions[ws['A1'].row].height = 30.0
	for col in ws['B:AG']:
	     ws.column_dimensions[col[0].column].width = 4.0

	ws["AI3"] = "B1"
	ws["AJ3"] = "B2"
	ws["AK3"] = "B3"
	ws["AL3"] = "B4"
	ws["AM3"] = "Q1"
	ws["AN3"] = "Q2"
	for i in range(4,11):
		ws["AI"+str(i)] = '=CONT.SE(B'+str(i)+':AG'+str(i)+'; "B1") '
		ws["AJ"+str(i)] = '=CONT.SE(B'+str(i)+':AG'+str(i)+'; "B2") '
		ws["AK"+str(i)] = '=CONT.SE(B'+str(i)+':AG'+str(i)+'; "B3") '
		ws["AL"+str(i)] = '=CONT.SE(B'+str(i)+':AG'+str(i)+'; "B4") '
		ws["AM"+str(i)] = '=CONT.SE(B'+str(i)+':AG'+str(i)+'; "Q1") '
		ws["AN"+str(i)] = '=CONT.SE(B'+str(i)+':AG'+str(i)+'; "Q2") '

	# Salvando
	nome_arquivo = escala[0][0]+'.xlsx'
	wb.save(filename=nome_arquivo)

def le_dados():
	data['estacao'] = input("nome estacao: ")
	data['dia_inicio'] = int(input("dia da mudança: "))
	data['func_num'] = int(input("numero de funcionarios: "))
	data['mes'] = input("mes: ")

	func_nomes = []
	func_Ps = []
	for i in range(data['func_num']):
		func_nomes.append(input("nome: "))
		func_Ps.append(int(input("P: ")))

	data['func_nomes'] = func_nomes
	data['func_Ps'] = func_Ps


# Importa as definições de setup.dat
def le_dados_auto():
	with open(sys.argv[1], "r") as entrada:
		print (sys.argv[1])
		data['estacao'] = entrada.readline()[:-1]
		data['mes'] = entrada.readline()[:-1]
		data['dia_inicio'] = int(entrada.readline())
		data['func_nomes'] = entrada.readline()[:-1].split(',')

def gera_tabela():
	# Titulo
	escala.append(["Escala ASO1 - " + data['estacao'] + " - " + data['mes'], ""])

	dias = bd['dias_mes'][data['mes']]

	# Sequencia de dias
	lista_dias = ["", "Dias"]
	for d in range(dias):
		lista_dias.append((data['dia_inicio']+d-1)%dias+1)
	escala.append(lista_dias)

	# Sequencia de dias da semana
	lista_sem = ["", "Ps"]
	for d in range(dias):
		lista_sem.append(bd['semana'][d%7])
	escala.append(lista_sem)

	distrib = preenche()
	postos = []
	c = 0
	for f in range(len(data['func_nomes'])):
		p = []
		f_nome = data['func_nomes'][f]
		p.append(f_nome)
		p.append(bd['pessoal'][f_nome])

		for i in range(dias):
			if distrib[c][i] == 0:
				p.append("")
			elif distrib[c][i] == 1:
				p.append("F")
			elif (distrib[c][i]%2)==0:
				p.append("B"+str(int(distrib[c][i]/2)))
			else:
				p.append("Q"+str(int(distrib[c][i]/2)))

		c += 1
		escala.append(p)


	print ("Arquivo gerado no diretorio: ")

def preenche():
	nomes = data['func_nomes']
	func = len(nomes)
	func_fixos = int(func/2)
	dias = bd['dias_mes'][data['mes']]

	# postos_trab = np.zeros((func,func_fixos))

	dist_postos = np.zeros((func,dias))

	# Atribui as folgas do funcionario
	f=0
	for n in nomes:
		p = bd['pessoal'][n]
		print(p, bd['folgas'][p])
		folgas = bd['folgas'][p]
		for d in folgas:
			dist_postos[f][d] = 1
		f += 1

	# Cria a sequencia de postos a trabalhar
	postos = bd['estacao'][data['estacao']]['postos']

	print (postos)

	pula = 0
	posto = 0
	for f in range(func_fixos):
		n_folgas = 0
		for d in range(dias):
			if dist_postos[f][d] == 1:
				n_folgas += 1
				pula += 1
				print(f+d-n_folgas+pula)
				dist_postos[f+func_fixos][d] = postos[(f+d-n_folgas+pula)%len(postos)]
			else:
				pula = 0
				print(f+d-n_folgas)
				dist_postos[f][d] = postos[(f+d-n_folgas)%len(postos)]

	print (dist_postos)
	return dist_postos



'''                    Programa                           '''


with open("data/data.json", "r") as read_file:
	bd = json.load(read_file)

# print_dic (bd)
escala = []
data = {}

le_dados_auto()

gera_tabela()

gera_xls()
